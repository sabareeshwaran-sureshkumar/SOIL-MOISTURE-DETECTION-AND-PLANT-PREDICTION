from sklearn.tree import DecisionTreeClassifier
import numpy as np
# import pandas as pd
import streamlit as st
import openpyxl
from collections.abc import MutableMapping


# Define the training data
soil_moisture = np.array([30, 60, 80, 20, 40, 70, 50, 90, 10, 60, 80, 20, 40, 70, 50, 90, 10, 60, 80, 20, 40, 70, 50, 90, 10, 60, 80, 20, 40, 70, 50, 90, 10, 60, 80, 20, 40, 70, 50, 90, 10, 60, 80, 20, 40, 70, 50, 90])
plant_type = np.array(['Tomatoes', 'Sunflowers', 'Beans', 'Radishes', 'Lettuce', 'Cucumbers', 'Peppers', 'Pumpkins', 'Carrots', 'Spinach', 'Squash', 'Onions', 'Broccoli', 'Corn', 'Watermelon', 'Cabbage', 'Beets', 'Eggplant', 'Okra', 'Garlic', 'Kale', 'Zucchini', 'Strawberries', 'Cauliflower', 'Celery', 'Peas', 'Artichokes', 'Asparagus', 'Brussels sprouts', 'Potatoes', 'Rhubarb', 'Sweet potatoes', 'Tarragon', 'Rosemary', 'Cilantro', 'Chives', 'Dill', 'Parsley', 'Sage', 'Thyme','Mint', 'Oregano', 'Basil', 'Lemon balm', 'Catnip', 'Lavender', 'Marjoram', 'Bee balm'])

# Create an instance of the DecisionTreeClassifier algorithm
clf = DecisionTreeClassifier(random_state=42)

# Fit the algorithm to the training data
clf.fit(soil_moisture.reshape(-1, 1), plant_type)

# Load the Excel workbook
workbook = openpyxl.load_workbook(r'C:\Users\sabar\OneDrive\Desktop\machinelearning\Book2.xlsx')

# Select the sheet you want to read from
sheet = workbook['Table 0']

# Get the last row and column with data
max_row1 = sheet.max_row
max_column = sheet.max_column

a=[]
for i in range(max_row1,max_row1+1):
    e15=sheet['E'+str(i)]
    a.append(e15.value)
    i+=1
print(a)
num=int(a[0])

# Define a new data point to predict the plant type
new_data = np.array([num]).reshape(-1, 1)

# Use the trained algorithm to predict the plant type of the new data point
predicted_plant_type = clf.predict(new_data)
st.title("Your Soil Moisture Value is :"+str(num))
st.write('The best fitted plant type for the soil moisture: <span style="color:red;">', predicted_plant_type[0], '</span>', unsafe_allow_html=True)



 
